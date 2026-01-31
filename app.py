from flask import Flask, render_template, jsonify, request, send_file
from flask_cors import CORS
import cv2
import numpy as np
import face_recognition
import os
from datetime import datetime
import openpyxl
from openpyxl import Workbook
import time
import json

app = Flask(__name__)
CORS(app)


class SmartAttendanceBackend:
    def __init__(self):
        self.dataset_path = 'dataset'
        self.attendance_file = 'attendance.xlsx'
        self.model_file = 'trained_model.npz'
        self.encodings = []
        self.names = []
        self.known_faces = []
        self.known_names = []
        self.marked_today = set()

    def create_dataset_folder(self):
        if not os.path.exists(self.dataset_path):
            os.makedirs(self.dataset_path)
        return True

    def capture_student_photos(self, student_name, num_photos=10):
        """Capture photos with real-time preview and face detection feedback"""
        self.create_dataset_folder()
        student_folder = os.path.join(self.dataset_path, student_name)

        if not os.path.exists(student_folder):
            os.makedirs(student_folder)

        existing_photos = len([f for f in os.listdir(student_folder)
                              if f.lower().endswith(('.jpg', '.jpeg', '.png'))])

        cap = cv2.VideoCapture(0)
        if not cap.isOpened():
            cap = cv2.VideoCapture(1)

        if not cap.isOpened():
            yield {'status': 'error', 'message': 'Cannot open camera'}
            return

        # Set camera resolution for better quality
        cap.set(cv2.CAP_PROP_FRAME_WIDTH, 640)
        cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 480)
        cap.set(cv2.CAP_PROP_FPS, 30)

        photos_captured = 0
        start_number = existing_photos + 1
        frame_count = 0
        face_detected = False

        print(f"\n📸 Capturing photos for: {student_name}")
        print(f"   Position yourself in front of camera with good lighting")
        print(f"   Need {num_photos} photos with clear face visibility")

        while photos_captured < num_photos:
            ret, frame = cap.read()
            if not ret:
                continue

            frame_count += 1
            
            # Detect faces in current frame
            rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            face_locations = face_recognition.face_locations(rgb_frame, model='hog')
            
            # Draw face rectangles for preview
            for top, right, bottom, left in face_locations:
                cv2.rectangle(frame, (left, top), (right, bottom), (0, 255, 0), 2)
                cv2.rectangle(frame, (left, top-35), (right, top), (0, 255, 0), cv2.FILLED)
                cv2.putText(frame, f'Face {photos_captured+1}/{num_photos}', 
                           (left+6, top-6), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255,255,255), 1)
            
            # Add instruction text
            cv2.putText(frame, f'Captured: {photos_captured}/{num_photos}', 
                       (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2)
            
            if len(face_locations) > 0:
                cv2.putText(frame, 'Face DETECTED - Good position!', 
                           (10, 70), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2)
                face_detected = True
            else:
                cv2.putText(frame, 'No face detected - Check lighting/angle', 
                           (10, 70), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 0, 255), 2)
                face_detected = False

            cv2.imshow(f'Capturing: {student_name}', frame)

            # Only save if face detected and every other frame (to get variation)
            if face_detected and frame_count % 2 == 0:
                photo_name = f"photo_{start_number + photos_captured}.jpg"
                photo_path = os.path.join(student_folder, photo_name)
                cv2.imwrite(photo_path, frame)
                photos_captured += 1
                print(f"   ✓ Saved photo {photos_captured}/{num_photos}")

                yield {
                    'status': 'capturing',
                    'current': photos_captured,
                    'total': num_photos,
                    'student': student_name,
                    'message': f'Photo {photos_captured} saved with face detected'
                }

                time.sleep(0.3)

            if cv2.waitKey(1) & 0xFF == ord('q'):
                break

        cap.release()
        cv2.destroyAllWindows()

        if photos_captured < num_photos:
            yield {
                'status': 'error',
                'message': f'Only {photos_captured} photos captured (need {num_photos}). Check lighting!'
            }
        else:
            print(f"✓ Captured all {photos_captured} photos successfully\n")
            yield {
                'status': 'complete',
                'photos_captured': photos_captured,
                'student': student_name,
                'path': student_folder
            }

    def train_faces(self):
        """Train with improved parameters"""
        if not os.path.exists(self.dataset_path):
            yield {'status': 'error', 'message': 'Dataset folder not found'}
            return

        person_folders = [f for f in os.listdir(self.dataset_path)
                         if os.path.isdir(os.path.join(self.dataset_path, f))]

        if not person_folders:
            yield {'status': 'error', 'message': 'No students in dataset'}
            return

        self.encodings = []
        self.names = []

        print("\n🧠 Training face recognition model...")

        for person_name in person_folders:
            person_path = os.path.join(self.dataset_path, person_name)
            image_files = [f for f in os.listdir(person_path)
                          if f.lower().endswith(('.jpg', '.jpeg', '.png'))]

            if not image_files:
                continue

            print(f"   Processing: {person_name} ({len(image_files)} photos)")
            yield {'status': 'processing', 'person': person_name, 'images': len(image_files)}

            successful_encodings = 0
            for img_file in image_files:
                img_path = os.path.join(person_path, img_file)
                try:
                    image = face_recognition.load_image_file(img_path)
                    
                    # Use HOG model for better accuracy (slower but more reliable)
                    face_locations = face_recognition.face_locations(image, model='hog')
                    
                    if face_locations:
                        # Get encoding for first face found
                        face_encs = face_recognition.face_encodings(image, face_locations, num_jitters=2)
                        
                        if face_encs:
                            encoding = np.array(face_encs[0], dtype=np.float64)
                            self.encodings.append(encoding)
                            self.names.append(person_name)
                            successful_encodings += 1
                    else:
                        print(f"      ⚠ No face in {img_file}")
                except Exception as e:
                    print(f"      ⚠ Error processing {img_file}: {e}")

            print(f"   ✓ {successful_encodings}/{len(image_files)} photos encoded successfully")

        if not self.encodings:
            print("❌ ERROR: No faces were successfully encoded!")
            print("   Solution: Ensure photos have clear, well-lit faces")
            yield {'status': 'error', 'message': 'No faces encoded. Check photo quality!'}
            return

        # Save model
        encodings_array = np.array(self.encodings, dtype=np.float64)
        names_array = np.array(self.names, dtype=object)
        np.savez_compressed(self.model_file, encodings=encodings_array, names=names_array)

        print(f"✓ Model trained with {len(self.encodings)} face encodings\n")
        yield {
            'status': 'complete',
            'total_faces': len(self.encodings),
            'unique_people': len(set(self.names)),
            'people_breakdown': {person: self.names.count(person) for person in set(self.names)}
        }

    def load_known_faces(self):
        if not os.path.exists(self.model_file):
            return False
        try:
            data = np.load(self.model_file, allow_pickle=True)
            encodings = data['encodings']
            names = data['names']
            self.known_faces = [np.array(enc, dtype=np.float64) for enc in encodings]
            self.known_names = [str(n) for n in names]
            print(f"✓ Loaded model with {len(self.known_faces)} encodings")
            return True
        except Exception as e:
            print(f"❌ Error loading model: {e}")
            return False

    def init_excel(self):
        if not os.path.exists(self.attendance_file):
            wb = Workbook()
            ws = wb.active
            ws.title = "Attendance"
            ws['A1'] = "Name"
            ws['B1'] = "Date"
            ws['C1'] = "Time"
            ws['D1'] = "Status"
            from openpyxl.styles import Font
            for cell in ws[1]:
                cell.font = Font(bold=True)
            wb.save(self.attendance_file)

    def mark_attendance(self, name):
        today = datetime.now().strftime('%Y-%m-%d')
        key = (name, today)
        if key in self.marked_today:
            return False

        self.init_excel()
        try:
            wb = openpyxl.load_workbook(self.attendance_file)
            ws = wb.active
            now = datetime.now()
            ws.append([name, now.strftime('%Y-%m-%d'), now.strftime('%H:%M:%S'), "Present"])
            wb.save(self.attendance_file)
            self.marked_today.add(key)
            return True
        except Exception:
            return False

    def get_attendance_records(self):
        if not os.path.exists(self.attendance_file):
            return []
        try:
            wb = openpyxl.load_workbook(self.attendance_file)
            ws = wb.active
            records = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:
                    records.append({
                        'name': row[0],
                        'date': str(row[1]),
                        'time': str(row[2]),
                        'status': row[3]
                    })
            return records
        except Exception:
            return []

    def get_students_list(self):
        if not os.path.exists(self.dataset_path):
            return []
        students = []
        for folder in os.listdir(self.dataset_path):
            folder_path = os.path.join(self.dataset_path, folder)
            if os.path.isdir(folder_path):
                photo_count = len([f for f in os.listdir(folder_path)
                                  if f.lower().endswith(('.jpg', '.jpeg', '.png'))])
                students.append({'name': folder, 'photos': photo_count})
        return students

    def run_attendance_loop(self):
        """Run face recognition with improved detection"""
        print("\n👁️ STARTING ATTENDANCE MODE")
        
        if not self.load_known_faces():
            raise Exception("Model not trained. Please train first!")

        cap = cv2.VideoCapture(0)
        if not cap.isOpened():
            cap = cv2.VideoCapture(1)
        if not cap.isOpened():
            raise Exception("Cannot open camera")

        # Optimize camera
        cap.set(cv2.CAP_PROP_FRAME_WIDTH, 640)
        cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 480)
        cap.set(cv2.CAP_PROP_FPS, 30)

        self.init_excel()
        cv2.namedWindow("ATTENDANCE", cv2.WINDOW_NORMAL)
        cv2.resizeWindow("ATTENDANCE", 800, 600)
        print("Camera opened. Press 'q' to stop.\n")

        frame_count = 0
        detected_today = {}  # Track detections in this session

        while True:
            ret, frame = cap.read()
            if not ret:
                break

            frame_count += 1

            # Process every 2nd frame for speed
            if frame_count % 2 == 0:
                small = cv2.resize(frame, (0, 0), fx=0.5, fy=0.5)
                rgb_small = cv2.cvtColor(small, cv2.COLOR_BGR2RGB)

                try:
                    # Detect faces with HOG (more reliable)
                    face_locations_small = face_recognition.face_locations(rgb_small, model='hog')
                    
                    if face_locations_small:
                        # Scale back to full frame
                        face_locations = [(t*2, r*2, b*2, l*2) for t, r, b, l in face_locations_small]
                        
                        # Encode on full frame for accuracy
                        rgb_full = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                        face_encodings = face_recognition.face_encodings(rgb_full, face_locations, num_jitters=1)

                        for face_enc, loc in zip(face_encodings, face_locations):
                            if not self.known_faces:
                                continue

                            distances = face_recognition.face_distance(self.known_faces, face_enc)
                            best_idx = int(np.argmin(distances))
                            best_distance = distances[best_idx]
                            
                            # Stricter threshold for better accuracy
                            if best_distance < 0.55:
                                name = self.known_names[best_idx]
                                confidence = (1 - best_distance) * 100
                                
                                # Mark attendance
                                if self.mark_attendance(name):
                                    print(f"✓ {name} marked present (confidence: {confidence:.1f}%)")
                                    detected_today[name] = detected_today.get(name, 0) + 1

                                top, right, bottom, left = loc
                                # Green box for recognized face
                                cv2.rectangle(frame, (left, top), (right, bottom), (0, 255, 0), 2)
                                cv2.rectangle(frame, (left, bottom-35), (right, bottom), (0, 255, 0), cv2.FILLED)
                                cv2.putText(frame, f'{name} ({confidence:.0f}%)', (left+6, bottom-6),
                                           cv2.FONT_HERSHEY_SIMPLEX, 0.8, (255, 255, 255), 2)
                            else:
                                # Red box for unknown face
                                top, right, bottom, left = loc
                                cv2.rectangle(frame, (left, top), (right, bottom), (0, 0, 255), 2)
                                cv2.putText(frame, f'Unknown ({1-best_distance:.0%})', (left+6, top-6),
                                           cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 0, 255), 2)

                except Exception as e:
                    print(f"Error in face recognition: {e}")

            # Show stats
            cv2.putText(frame, f'Detected: {len(detected_today)} students', 
                       (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2)
            cv2.putText(frame, 'Press q to stop', 
                       (10, 470), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 255), 1)

            cv2.imshow("ATTENDANCE", frame)
            if cv2.waitKey(1) & 0xFF == ord('q'):
                break

        cap.release()
        cv2.destroyAllWindows()
        print(f"\n✓ Session ended. {len(detected_today)} students detected.\n")


backend = SmartAttendanceBackend()


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/api/status')
def get_status():
    camera_available = False
    try:
        cap = cv2.VideoCapture(0)
        if cap.isOpened():
            camera_available = True
            cap.release()
    except:
        pass

    return jsonify({
        'model_trained': os.path.exists(backend.model_file),
        'students': backend.get_students_list(),
        'camera_available': camera_available
    })


@app.route('/api/capture-student', methods=['POST'])
def capture_student():
    data = request.json
    student_name = data.get('name', '').strip()
    num_photos = int(data.get('photos', 10))

    if not student_name:
        return jsonify({'error': 'Student name required'}), 400

    def generate():
        for update in backend.capture_student_photos(student_name, num_photos):
            yield 'data: ' + json.dumps(update) + '\n\n'

    return app.response_class(generate(), mimetype='text/event-stream')


@app.route('/api/train-model', methods=['POST'])
def train_model():
    def generate():
        for update in backend.train_faces():
            yield 'data: ' + json.dumps(update) + '\n\n'

    return app.response_class(generate(), mimetype='text/event-stream')


@app.route('/api/attendance-records')
def get_records():
    return jsonify(backend.get_attendance_records())


@app.route('/api/download-attendance')
def download_attendance():
    if os.path.exists(backend.attendance_file):
        return send_file(backend.attendance_file, as_attachment=True)
    return jsonify({'error': 'No file found'}), 404


@app.route('/api/students')
def get_students():
    return jsonify(backend.get_students_list())


@app.route('/api/start-attendance', methods=['POST'])
def start_attendance():
    try:
        backend.run_attendance_loop()
        return jsonify({'message': 'Attendance finished', 'status': 'success'})
    except Exception as e:
        print(f"Error: {e}")
        return jsonify({'error': str(e), 'status': 'failed'}), 500


if __name__ == '__main__':
    app.run(debug=True, port=5000)
